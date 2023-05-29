import openpyxl
wb = openpyxl.load_workbook('main.xlsx')
sheet = wb['USER']

def find_data(desired_mobile,desired_password,sheet = sheet):
    for row in sheet.iter_rows(min_row=1): 
        mobile = row[3].value
        password = row[4].value
        if mobile == desired_mobile and password == desired_password:
            print("LOGIN SUCCESS")
            print(row[0].value)
            return row[0].value
            break  
    else:
        print("ERROR DATA NOT FOUND")
        

def insert_data(data_to_append,sheet=sheet):
    data_to_append = [sheet.max_row+1] + data_to_append
    sheet.append(data_to_append)
    wb.save('main.xlsx')

def login():
    print("----LOGIN PAGE----")
    u_name = input("ENTER USER NAME : ")
    u_mo_number = int(input("ENTER USER PHONE NUMBER : "))
    passwd = input("ENTER USER PASSWORD : ")
    return find_data(u_mo_number,passwd)
    

def signin():
    print("----SIGN IN PAGE----") 
    u_name = input("ENTER USER NAME : ")
    u_email = input("ENTER USER EMAIL : ")
    u_mo_number = int(input("ENTER USER PHONE NUMBER : "))
    u_passwd = input("SET NEW PASSWORD : ")
    data_to_append = [u_name,u_email,u_mo_number,u_passwd]
    insert_data(data_to_append)

def user():
    s_l = int(input("SIGN IN (0) OR LOGIN (1)"))
    if s_l == 1:
        return login()
    elif s_l == 0:
        signin()
    else:
        print("ERROR !")
        return


